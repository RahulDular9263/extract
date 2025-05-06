from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import fitz  # PyMuPDF
import pandas as pd
import io
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

app = Flask(__name__)
CORS(app)

def extract_comments(doc):
    comments = []
    for i, page in enumerate(doc):
        annots = page.annots()
        if annots:
            for annot in annots:
                if annot.type[1] in ["Text", "FreeText"]:
                    comments.append({
                        "Page": i + 1,
                        "Author": annot.info.get("title", ""),
                        "Comment": annot.info.get("content", "")
                    })
    return comments

@app.route('/comments', methods=['POST'])
def get_comments():
    file = request.files['pdf']
    doc = fitz.open(stream=file.read(), filetype="pdf")
    comments = extract_comments(doc)
    return jsonify({"comments": comments})

@app.route('/extract', methods=['POST'])
def generate_excel():
    file = request.files['pdf']
    doc = fitz.open(stream=file.read(), filetype="pdf")
    comments = extract_comments(doc)
    df = pd.DataFrame(comments)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Comments", startrow=0)
        ws = writer.sheets["Comments"]

        for i, column in enumerate(df.columns, 1):
            max_len = max(df[column].astype(str).map(len).max(), len(column)) + 2
            ws.column_dimensions[get_column_letter(i)].width = max_len
            ws.cell(row=1, column=i).font = Font(bold=True)

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="PDF_Comments_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == '__main__':
    app.run(debug=True)